# your_package/document_processor/table_data_handler.py
import logging, csv, aiofiles, os
from typing import Any
from .utils import clean_text

logger = logging.getLogger("document-processor")

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except Exception:
    XLRD_AVAILABLE = False

async def extract_text_from_excel(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    text = ""
    try:
        if ext == ".xlsx":
            if not OPENPYXL_AVAILABLE:
                raise Exception("openpyxl이 설치되어야 .xlsx 파일을 처리할 수 있습니다.")
            wb = load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text += f"\n=== 시트: {sheet_name} ===\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(v) for v in row if v is not None)
                    if row_text.strip():
                        text += row_text + "\n"
        elif ext == ".xls":
            if not XLRD_AVAILABLE:
                raise Exception("xlrd가 설치되어야 .xls 파일을 처리할 수 있습니다.")
            wb = xlrd.open_workbook(file_path)
            for sheet in wb.sheets():
                text += f"\n=== 시트: {sheet.name} ===\n"
                for row_idx in range(sheet.nrows):
                    row_values = sheet.row_values(row_idx)
                    row_text = " | ".join(str(v) for v in row_values if v)
                    if row_text.strip():
                        text += row_text + "\n"
        else:
            raise Exception(f"지원하지 않는 Excel 형식입니다: {ext}")
        return clean_text(text)
    except Exception as e:
        logger.error(f"Error extracting text from Excel {file_path}: {e}")
        raise

async def extract_text_from_csv(file_path: str, encoding: str = "utf-8") -> str:
    text = ""
    encodings_to_try = [encoding, "cp949", "euc-kr", "utf-8-sig", "latin-1"]
    
    for enc in encodings_to_try:
        try:
            async with aiofiles.open(file_path, mode="r", encoding=enc) as f:
                content = await f.read()
                reader = csv.reader(content.splitlines())
                for row in reader:
                    row_text = " | ".join(row)
                    if row_text.strip():
                        text += row_text + "\n"
                logger.info(f"Successfully read CSV with encoding: {enc}")
                return clean_text(text)
        except UnicodeDecodeError:
            logger.debug(f"Failed to decode with {enc}, trying next encoding...")
            continue
        except Exception as e:
            logger.error(f"Error extracting text from CSV {file_path} with {enc}: {e}")
            if enc == encodings_to_try[-1]:  # 마지막 시도였다면
                raise
    
    raise Exception(f"Could not decode CSV file with any of the attempted encodings: {encodings_to_try}")